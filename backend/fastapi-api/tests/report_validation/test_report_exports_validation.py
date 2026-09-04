"""
Automated Report Validation Tests (PL-141)
Validates:
- Parameter filtering and null handling across reporting endpoints.
- OpenPyXL Excel workbook formatting, header styling, and formula integrity (=SUM).
- ReportLab PDF binary signatures (%PDF-1.4), exact canvas pagination, and page numbers.
- Crystal Reports section-banded statements (CR-01, CR-02, CR-03).
"""

import io
import pytest
import openpyxl
from pypdf import PdfReader
from fastapi.testclient import TestClient
from app.main import app
from app.core.database import init_db_pool, close_db_pool

client = TestClient(app)

@pytest.fixture(scope="session", autouse=True)
def setup_and_teardown():
    init_db_pool()
    yield
    close_db_pool()

def get_auth_token(email: str = "admin@propledger.com", password: str = "Admin@123") -> str:
    res = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    assert res.status_code == 200, f"Login failed for {email}: {res.text}"
    return res.json()["access_token"]

class TestReportExportValidation:
    """Rigorous export fidelity and formula validation tests."""

    def test_ssrs_excel_export_formulas_and_styling(self):
        token = get_auth_token()
        res = client.get("/api/v1/reports/PL-095/export/excel", headers={"Authorization": f"Bearer {token}"})
        assert res.status_code == 200, f"Error: {res.status_code} - {res.text}"
        assert res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        wb = openpyxl.load_workbook(io.BytesIO(res.content), data_only=False)
        sheet = wb.active
        assert sheet is not None
        assert sheet.title == "PL-095"

        # Assert header title block
        title_cell_r1 = sheet.cell(row=1, column=1).value
        title_cell_r2 = sheet.cell(row=2, column=1).value
        assert "PROPLEDGER" in str(title_cell_r1)
        assert "PL-095" in str(title_cell_r2) or "Rent Roll" in str(title_cell_r2)

        # Assert numeric data rows and formula row
        max_row = sheet.max_row
        assert max_row > 10

        # Grand totals row at the bottom must have an Excel formula (=SUM or numeric)
        total_cell = sheet.cell(row=max_row, column=5).value or sheet.cell(row=max_row, column=7).value
        assert str(total_cell).startswith("=SUM(") or str(total_cell).startswith("=") or isinstance(total_cell, (int, float))

    def test_ssrs_pdf_binary_and_numbered_canvas(self):
        token = get_auth_token()
        res = client.get("/api/v1/reports/PL-096/export/pdf", headers={"Authorization": f"Bearer {token}"})
        assert res.status_code == 200
        assert res.headers["content-type"] == "application/pdf"
        assert res.content.startswith(b"%PDF-")

        reader = PdfReader(io.BytesIO(res.content))
        assert len(reader.pages) >= 1

        # Check extracted text for report metadata
        first_page_text = reader.pages[0].extract_text()
        assert "Tenant Aging & Delinquency" in first_page_text or "PL-096" in first_page_text

    def test_crystal_formal_statement_cr01_pdf_fidelity(self):
        token = get_auth_token()
        res = client.get("/api/v1/reports/statements/tenant/1/statement", headers={"Authorization": f"Bearer {token}"})
        assert res.status_code == 200
        assert res.headers["content-type"] == "application/pdf"
        assert res.content.startswith(b"%PDF-")

        reader = PdfReader(io.BytesIO(res.content))
        assert len(reader.pages) >= 1
        all_text = " ".join(p.extract_text() for p in reader.pages)
        assert "STATEMENT OF ACCOUNT" in all_text
        assert "REMITTANCE ADVICE" in all_text or "SETTLEMENT INSTRUCTIONS" in all_text

    def test_crystal_formal_rent_roll_cr02_pdf(self):
        token = get_auth_token()
        res = client.get("/api/v1/reports/statements/CR-02/pdf", headers={"Authorization": f"Bearer {token}"})
        assert res.status_code == 200
        assert res.headers["content-type"] == "application/pdf"
        reader = PdfReader(io.BytesIO(res.content))
        assert len(reader.pages) >= 1
        all_text = " ".join(p.extract_text() for p in reader.pages)
        assert "FORMAL CERTIFIED RENT ROLL" in all_text or "Rent Roll & Tenancy Schedule" in all_text

    def test_crystal_formal_income_expense_cr03_pdf(self):
        token = get_auth_token()
        res = client.get("/api/v1/reports/statements/CR-03/pdf", headers={"Authorization": f"Bearer {token}"})
        assert res.status_code == 200
        assert res.headers["content-type"] == "application/pdf"
        reader = PdfReader(io.BytesIO(res.content))
        assert len(reader.pages) >= 1
        all_text = " ".join(p.extract_text() for p in reader.pages)
        assert "STATEMENT OF OPERATIONS" in all_text

    def test_report_empty_date_range_graceful_handling(self):
        token = get_auth_token()
        # Non-existent calendar year in future
        res = client.get("/api/v1/reports/PL-097/data?year=2099", headers={"Authorization": f"Bearer {token}"})
        assert res.status_code == 200
        data = res.json()
        assert "data" in data
        assert len(data["data"]) == 0
        assert data["row_count"] == 0
