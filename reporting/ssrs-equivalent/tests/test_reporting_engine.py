"""
Automated Test Suite for PropLedger SSRS-Equivalent Enterprise Reporting Engine.
Tests all 14 report families (PL-095 through PL-108), OpenPyXL Excel rendering,
ReportLab PDF rendering, parameter validation, and registry discovery.
"""
import io
import os
import sys
from pathlib import Path
import pytest
import openpyxl

# Add ssrs-equivalent directory to sys.path
REPORTING_DIR = Path(__file__).resolve().parent.parent
if str(REPORTING_DIR) not in sys.path:
    sys.path.insert(0, str(REPORTING_DIR))

from registry import ReportRegistry
from reports import REPORT_CLASSES


EXPECTED_REPORT_CODES = [
    "PL-095", "PL-096", "PL-097", "PL-098", "PL-099",
    "PL-100", "PL-101", "PL-102", "PL-103", "PL-104",
    "PL-105", "PL-106", "PL-107", "PL-108"
]


class TestReportRegistry:
    """Tests for report registry, metadata discovery, and schema definitions."""

    def test_all_fourteen_reports_registered(self):
        """Verify all 14 PRD Part Q report families are registered."""
        ReportRegistry.initialize()
        reports = ReportRegistry.list_reports()
        registered_codes = [r["report_code"] for r in reports]
        for code in EXPECTED_REPORT_CODES:
            assert code in registered_codes, f"Report {code} missing from registry"
        assert len(reports) >= 14

    @pytest.mark.parametrize("report_code", EXPECTED_REPORT_CODES)
    def test_report_metadata_structure(self, report_code):
        """Verify each report has complete, valid metadata definition."""
        report = ReportRegistry.get_report(report_code)
        assert report is not None, f"Report {report_code} failed to instantiate"
        meta = report.get_metadata()
        assert meta["report_code"] == report_code
        assert len(meta["title"]) > 0
        assert len(meta["category"]) > 0
        assert len(meta["description"]) > 0
        assert len(meta["columns"]) >= 5
        assert isinstance(meta["parameters"], dict)

        # Validate column definitions
        for col in meta["columns"]:
            assert "key" in col
            assert "label" in col
            assert "type" in col
            assert col["type"] in ("string", "currency", "number", "percent", "date")


class TestReportExecutionAndExport:
    """Tests data fetching, Excel binary generation, and PDF binary generation."""

    @pytest.mark.parametrize("report_code", EXPECTED_REPORT_CODES)
    def test_fetch_data_from_database(self, report_code):
        """Verify report queries real PostgreSQL database without errors."""
        report = ReportRegistry.get_report(report_code)
        data = report.fetch_data({"limit": 50})
        assert isinstance(data, list)
        if len(data) > 0:
            first_row = data[0]
            # Ensure key columns from definition exist in output
            expected_first_col = report.columns[0]["key"]
            assert expected_first_col in first_row, (
                f"Report {report_code} missing expected column '{expected_first_col}' in query result"
            )

    @pytest.mark.parametrize("report_code", EXPECTED_REPORT_CODES)
    def test_export_excel_generation(self, report_code):
        """Verify real binary .xlsx generation with OpenPyXL validity checks."""
        report = ReportRegistry.get_report(report_code)
        excel_bytes = report.export_excel({"limit": 20})
        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 2000, f"Excel file too small ({len(excel_bytes)} bytes)"

        # Load with openpyxl to verify workbook integrity
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        sheet = wb.active
        assert sheet.views.sheetView[0].showGridLines is True
        # Verify brand title and report title exist in title block
        assert "PROPLEDGER" in str(sheet.cell(row=1, column=1).value)
        assert report_code in str(sheet.cell(row=2, column=1).value)

    @pytest.mark.parametrize("report_code", EXPECTED_REPORT_CODES)
    def test_export_pdf_generation(self, report_code):
        """Verify real binary .pdf generation with PDF-1.x magic signature."""
        report = ReportRegistry.get_report(report_code)
        pdf_bytes = report.export_pdf({"limit": 20})
        assert isinstance(pdf_bytes, bytes)
        assert len(pdf_bytes) > 2000, f"PDF file too small ({len(pdf_bytes)} bytes)"
        # Standard PDF magic header
        assert pdf_bytes.startswith(b"%PDF-"), "Generated file is not a valid PDF document"

    def test_parameter_filtering(self):
        """Test parameter validation and filtering on RentRoll (PL-095)."""
        report = ReportRegistry.get_report("PL-095")
        filtered_data = report.fetch_data({"property_id": 1, "occupancy_status": "OCCUPIED", "limit": 100})
        assert isinstance(filtered_data, list)
        for row in filtered_data:
            assert row.get("occupancy_status") == "OCCUPIED"

    def test_summary_kpi_cards(self):
        """Test KPI summary cards calculation on Executive Dashboard (PL-108)."""
        report = ReportRegistry.get_report("PL-108")
        data = report.fetch_data({"limit": 50})
        kpis = report.get_summary_stats(data)
        assert len(kpis) >= 3
        labels = [k["label"] for k in kpis]
        assert "Total Units" in labels or "Portfolio Properties" in labels
