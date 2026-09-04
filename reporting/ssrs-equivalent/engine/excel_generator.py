"""
OpenPyXL Excel Generation Engine for PropLedger.
Produces styled corporate workbooks with dark navy headers, zebra striping,
frozen panes, auto-fit columns, and dynamic formula totals.
"""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """
    Renders report data into publication-quality styled Excel workbooks (.xlsx).
    """

    NAVY_HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    WHITE_BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    TOTALS_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    KPI_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    THIN_BORDER = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    TOTAL_BORDER = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )

    def __init__(self, report: Any):
        self.report = report

    def generate(
        self,
        data: List[Dict[str, Any]],
        params: Dict[str, Any],
        output_path: Optional[str] = None,
    ) -> bytes:
        """Generates formatted Excel workbook and returns binary bytes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.report.report_code[:31]  # Excel tab name max 31 chars
        ws.views.sheetView[0].showGridLines = True

        current_row = 1

        # 1. Company Brand Header
        ws.cell(row=current_row, column=1, value="PROPLEDGER ENTERPRISE PROPERTY MANAGEMENT & ANALYTICS")
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=10, bold=True, color="64748B")
        current_row += 1

        # 2. Report Title
        ws.cell(row=current_row, column=1, value=f"{self.report.report_code} — {self.report.title}")
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=15, bold=True, color="1E3A8A")
        current_row += 1

        # 3. Metadata Subtitle (Generated date, category, parameters)
        param_str = ", ".join(f"{k}={v}" for k, v in params.items() if v is not None) or "All Records"
        meta_text = (
            f"Category: {self.report.category}  |  "
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
            f"Filters: [{param_str}]"
        )
        ws.cell(row=current_row, column=1, value=meta_text)
        ws.cell(row=current_row, column=1).font = Font(name="Calibri", size=9, italic=True, color="475569")
        current_row += 2  # Blank line before table or KPIs

        # 4. Optional KPI Summary Cards
        summary_stats = self.report.get_summary_stats(data)
        if summary_stats:
            kpi_col = 1
            for stat in summary_stats[:6]:  # Show top 6 KPIs
                label_cell = ws.cell(row=current_row, column=kpi_col, value=stat.get("label"))
                label_cell.font = Font(name="Calibri", size=8, color="64748B", bold=True)
                label_cell.fill = self.KPI_FILL
                label_cell.border = self.THIN_BORDER
                label_cell.alignment = Alignment(horizontal="center", vertical="center")

                val_cell = ws.cell(row=current_row + 1, column=kpi_col, value=stat.get("value"))
                val_cell.font = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
                val_cell.fill = self.KPI_FILL
                val_cell.border = self.THIN_BORDER
                val_cell.alignment = Alignment(horizontal="center", vertical="center")
                kpi_col += 1
            current_row += 3

        # 5. Table Header Row
        header_row = current_row
        columns = self.report.columns

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_def["label"])
            cell.fill = self.NAVY_HEADER_FILL
            cell.font = self.WHITE_BOLD_FONT
            cell.border = self.THIN_BORDER
            align = col_def.get("align", "left")
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)

        ws.row_dimensions[header_row].height = 24
        current_row += 1
        data_start_row = current_row

        # Freeze panes below header
        ws.freeze_panes = f"A{data_start_row}"

        # 6. Data Rows
        for row_idx, record in enumerate(data):
            row_num = current_row
            is_zebra = (row_idx % 2 == 1)
            row_fill = self.ZEBRA_FILL if is_zebra else None

            for col_idx, col_def in enumerate(columns, start=1):
                key = col_def["key"]
                raw_val = record.get(key)
                col_type = col_def.get("type", "string")
                align = col_def.get("align", "left")

                cell = ws.cell(row=row_num, column=col_idx)
                if row_fill:
                    cell.fill = row_fill
                cell.border = self.THIN_BORDER

                if raw_val is None:
                    cell.value = "-"
                    cell.alignment = Alignment(horizontal="center")
                elif col_type == "currency":
                    try:
                        cell.value = float(raw_val)
                    except (ValueError, TypeError):
                        cell.value = raw_val
                    cell.number_format = '"₹"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_type == "percent":
                    try:
                        # Handle values passed as 85.5% vs 0.855
                        fval = float(raw_val)
                        if fval > 1.0:
                            fval = fval / 100.0
                        cell.value = fval
                    except (ValueError, TypeError):
                        cell.value = raw_val
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                elif col_type == "number":
                    try:
                        cell.value = float(raw_val) if "." in str(raw_val) else int(raw_val)
                    except (ValueError, TypeError):
                        cell.value = raw_val
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif col_type == "date":
                    cell.value = str(raw_val)
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.value = str(raw_val)
                    cell.alignment = Alignment(horizontal=align)

            ws.row_dimensions[row_num].height = 18
            current_row += 1

        data_end_row = current_row - 1

        # 7. Grand Total Row (if data exists and numeric/currency columns present)
        if data and data_end_row >= data_start_row:
            total_row = current_row
            ws.cell(row=total_row, column=1, value="Grand Total")
            ws.cell(row=total_row, column=1).font = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
            ws.cell(row=total_row, column=1).fill = self.TOTALS_FILL
            ws.cell(row=total_row, column=1).border = self.TOTAL_BORDER

            for col_idx in range(2, len(columns) + 1):
                col_def = columns[col_idx - 1]
                col_type = col_def.get("type", "string")
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(row=total_row, column=col_idx)
                cell.fill = self.TOTALS_FILL
                cell.border = self.TOTAL_BORDER

                if col_type in ("currency", "number"):
                    cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                    cell.font = Font(name="Calibri", size=10, bold=True)
                    if col_type == "currency":
                        cell.number_format = '"₹"#,##0.00'
                    else:
                        cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif col_type == "percent":
                    cell.value = f"=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                    cell.font = Font(name="Calibri", size=10, bold=True)
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = ""

            ws.row_dimensions[total_row].height = 20
            current_row += 1

        # 8. Auto-fit column widths
        for col_idx, col_def in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            # Default width based on column definition or calculated content length
            spec_width = col_def.get("width")
            if spec_width:
                ws.column_dimensions[col_letter].width = max(spec_width, 12)
            else:
                max_len = len(str(col_def["label"]))
                for r in range(data_start_row, min(data_start_row + 100, current_row)):
                    v = ws.cell(row=r, column=col_idx).value
                    if v:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        # 9. Save and return bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        binary_data = buffer.getvalue()

        if output_path:
            with open(output_path, "wb") as f:
                f.write(binary_data)

        return binary_data
